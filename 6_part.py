import csv
import math
import re
# import prettytable
import os
import datetime
import requests

from unittest import TestCase
import unittest

import numpy as np
import openpyxl as op
import pandas as pd
from jinja2 import Environment, FileSystemLoader
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter
from matplotlib import pyplot as plt
import matplotlib
import pdfkit
import multiprocessing


vacant_dic = {"name": "Название", "description": "Описание", "key_skills": "Навыки", "experience_id": "Опыт работы",
              "premium": "Премиум-вакансия", "employer_name": "Компания", "salary": "Оклад",
              "area_name": "Название региона", "published_at": "Дата публикации вакансии"}
currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}
formaterr_dict = {"AZN": "Манаты", "BYR": "Белорусские рубли",
                  "EUR": "Евро", "GEL": "Грузинский лари",
                  "KGS": "Киргизский сом", "KZT": "Тенге",
                  "RUR": "Рубли", "UAH": "Гривны",
                  "USD": "Доллары", "UZS": "Узбекский сум",
                  "noExperience": "Нет опыта",
                  "between1And3": "От 1 года до 3 лет",
                  "between3And6": "От 3 до 6 лет",
                  "moreThan6": "Более 6 лет",
                  "": "Нет значения"}

#Попытка мультипроцессинга №64
# def pool_work(self, x, dict_inYear_WithName, dict_inYear_WithName_salary):
#     time = x.published_at.split('-')
#     if not dict_inYear_WithName.__contains__(int(time[0])) and (x.name.__contains__(dataSet.job_name)
#                                                                 or x.name.__contains__(
#                 dataSet.job_name.lower())):
#         dict_inYear_WithName[int(time[0])] = 1
#         dict_inYear_WithName_salary[int(time[0])] = (float(x.salary.salary_from)
#                                                      * currency_to_rub[x.salary.salary_currency]
#                                                      + float(x.salary.salary_to)
#                                                      * currency_to_rub[x.salary.salary_currency]) \
#                                                     / 2
#     elif x.name.__contains__(dataSet.job_name) or x.name.__contains__(dataSet.job_name.lower()):
#         dict_inYear_WithName[int(time[0])] += 1
#         dict_inYear_WithName_salary[int(time[0])] += (float(x.salary.salary_from)
#                                                       * currency_to_rub[x.salary.salary_currency]
#                                                       + float(x.salary.salary_to)
#                                                       * currency_to_rub[x.salary.salary_currency]) \
#                                                      / 2
#     return (dict_inYear_WithName, dict_inYear_WithName_salary)
#
#
# def pool_work2(self, x, dict_inYear_noName, dict_inYear_noName_salary):
#     time = x.published_at.split('-')
#     if not dict_inYear_noName.__contains__(int(time[0])):
#         dict_inYear_noName[int(time[0])] = 1
#         dict_inYear_noName_salary[int(time[0])] = (float(x.salary.salary_from)
#                                                    * currency_to_rub[x.salary.salary_currency]
#                                                    + float(x.salary.salary_to)
#                                                    * currency_to_rub[x.salary.salary_currency]) \
#                                                   / 2
#     elif dict_inYear_noName.__contains__(int(time[0])):
#         dict_inYear_noName[int(time[0])] += 1
#         dict_inYear_noName_salary[int(time[0])] += (float(x.salary.salary_from)
#                                                     * currency_to_rub[x.salary.salary_currency]
#                                                     + float(x.salary.salary_to)
#                                                     * currency_to_rub[x.salary.salary_currency]) \
#                                                    / 2
#     return (dict_inYear_noName, dict_inYear_noName_salary)
#
#
# def pool_work3(self, x, temp_dict, temp_salary_dict):
#     city = x.area_name
#     if not temp_dict.__contains__(city):
#         temp_dict[city] = 1
#         temp_salary_dict[city] = (float(x.salary.salary_from)
#                                   * currency_to_rub[x.salary.salary_currency]
#                                   + float(x.salary.salary_to)
#                                   * currency_to_rub[x.salary.salary_currency]) \
#                                  / 2
#     elif temp_dict.__contains__(city):
#         temp_dict[city] += 1
#         temp_salary_dict[city] += (float(x.salary.salary_from)
#                                    * currency_to_rub[x.salary.salary_currency]
#                                    + float(x.salary.salary_to)
#                                    * currency_to_rub[x.salary.salary_currency]) \
#                                    / 2
#
#     return (temp_dict, temp_salary_dict)

curr_dict = {}

class InputConect():
    """
    Класс для обработки значений: фильтрация, сортировка и тд.

    Атрибуты:
        dataSet (DataSet) - Содержит в себе экземпляр обекта со всеми входными данными

        dict_inYear_noName (dict) - словарь для всех вакансий
        dict_inYear_noName_salary (dict) - словарь для заработной платы вакансий

        dict_inYear_WithName (dict) - словарь для вакансий определенной профессии
        dict_inYear_WithName_salary (dict) - словарь для заработной платы определенной профессии

        dict_inYear_City (dict) - словарь распределения всех вакансий по городам
        dict_inYear_City_salary (dict) - словарь распределения зарплат по городам
        temp_dict (dict) - временный словарь для обработки считываемых значений вакансий
        temp_salary_dict (dict) - временный словарь для обработки считываемых значений зарплаты
    """
    dataSet = ""
    dict_inYear_noName = {}
    dict_inYear_noName_salary = {}

    dict_inYear_WithName_and_city = {}
    dict_inYear_WithName_salary_and_city = {}

    dict_inYear_City = {}
    dict_inYear_City_salary = {}
    temp_dict = {}
    temp_salary_dict = {}

    def __init__(self, data, skip_init = False):
        """
        Инициализирует объект InputConect, путем назначения входного экземпляра DataSet полю InputConect

        :param data: (DataSet) - Содержит в себе экземпляр обекта со всеми входными данными

        """
        if not skip_init:
            self.dataSet = data

    def fill_title(self, filling_list, dic_naming, table):
        """
        Заполняет названия полей у таблицы table (prettytable) по макету значений вакансии

        :param filling_list (list) - лист вакансий, что считывается из словаря и далее передается в саму таблицу
        :param dic_naming (dict) - словарь со всеми вакансиями, в которой ключами служат значения из словаря макета вакансии - vacant_dic (dict)
        :param table (prettytable) - таблица для заполнения
        :return: таблица (prettytable)
        """
        global vacant_dic
        for text_info in vacant_dic:
            filling_list.append(dic_naming[text_info])
        if len(table.field_names) == 0:
            table.field_names = filling_list

    def prepare_vacancies(self, dataSet, dic_naming, table):
        """
        Подготовка вакансий и добавление их в таблицу и ее подготовка

        :param dataSet (DataSet) - Содержит в себе экземпляр обекта со всеми входными данными
        :param dic_naming (dict) - словарь со всеми вакансиями, в которой ключами служат значения из словаря макета вакансии - vacant_dic (dict)
        :param table (prettytable) - таблица для заполнения
        :return: таблица (prettytable)
        """
        global full_table_skills
        title_list = ["№"]
        self.fill_title(title_list, dic_naming, table)
        table.max_width = 20
        count = 0
        for row in dataSet.vacancies_objects:
            formated_row = self.formatter(row)
            for i in range(len(formated_row)):
                if i == 2:
                    temp_skills = '\n'.join(formated_row[i])
                    # Сторонний сбор данных для сортировки навыков
                    if not full_table_skills.__contains__(formated_row[0]):
                        full_table_skills[formated_row[0]] = []
                    if not list(full_table_skills[formated_row[0]]).__contains__(formated_row[i]):
                        full_table_skills[formated_row[0]].append(formated_row[i])

                    if len(temp_skills) > 100:
                        temp = list(temp_skills)
                        temp[100] = "..."
                        formated_row[i] = "".join(temp[0:101])
                    else:
                        formated_row[i] = temp_skills
                elif len(formated_row[i]) > 100:
                    temp = list(formated_row[i])
                    temp[100] = "..."
                    formated_row[i] = "".join(temp[0:101])
            count += 1
            table.add_row([count, formated_row[0], formated_row[1], formated_row[2], formated_row[3],
                           formated_row[4], formated_row[5], formated_row[6], formated_row[7], formated_row[8]])

    def formatter(self, rows):
        """
        Форматирование зарплаты, времени и стажа

        :param rows (list) - список из всех вакансий объекта DataSet
        :return: list из форматированных строк
        """
        global full_table_date
        global formaterr_dict
        result = []
        format_dic = {}
        salar_min = rows.salary.prepare_salary(str(int(float(rows.salary.salary_from))))
        salar_max = rows.salary.prepare_salary(str(int(float(rows.salary.salary_to))))
        if rows.salary.salary_gross == "Нет":
            rows.salary.salary_gross = "С вычетом налогов"
        elif rows.salary.salary_gross == "Да":
            rows.salary.salary_gross = "Без вычета налогов"
        salary_mean = "{0} - {1}".format(salar_min, str(salar_max))
        salary_answer = "{0} ({1}) ({2})".format(str(salary_mean), formaterr_dict[rows.salary.salary_currency],
                                                 rows.salary.salary_gross)
        # Перенос данных в правильную форму листа
        format_dic['name'] = rows.name
        format_dic['description'] = rows.description
        format_dic['key_skills'] = rows.key_skills
        format_dic['experience_id'] = formaterr_dict[rows.experience_id]
        format_dic['premium'] = rows.premium
        format_dic['employer_name'] = rows.employer_name
        format_dic['salary'] = salary_answer
        format_dic['area_name'] = rows.area_name

        pre_date = rows.published_at.split("T")[0].split("-")
        # Сторонний сбор данных для сортировки даты
        temp_date_time = rows.published_at.split("T")[1].split("+")[0]
        if not full_table_date.__contains__(format_dic["name"]):
            full_table_date[format_dic["name"]] = []
        if not list(full_table_date[format_dic["name"]]).__contains__(temp_date_time):
            full_table_date[format_dic["name"]].append((temp_date_time, format_dic["area_name"]))
        rows.published_at = "{0}.{1}.{2}".format(pre_date[2], pre_date[1], pre_date[0])
        format_dic['published_at'] = rows.published_at
        for item in format_dic:
            result.append(format_dic[item])
        return result

    def find_full_skills(self, x, full_skills):
        """
        Правило сортировки по всем скиллам. Осуществляет поиск полного списка навыков для каждой строчки (не укороченного)

        :param x (list) - список элементов одной вакансии
        :param full_skills (dict) - словарь полных навыков для каждой профессии
        :return: string содержащий все навыки для вакансии хранящейся в строке x
        """
        list_skills = full_skills[x[1]]
        redact_skills = x[3].replace(".", "")
        index = 0
        for i in range(len(list_skills)):
            str_skills = '\n'.join(list_skills[i])
            if str(str_skills).__contains__(redact_skills):
                index = i
                break
        return list_skills[index]

    def get_date_sort(self, x):
        """
        Правило сортировки по дате

        :param x (Vacancy) - объект Vacancy, содержащий элементы одной вакансии
        :return: string содержащий полную дату для вакансии хранящейся в строке x

        #>>> InputConect(DataSet()).get_date_sort(Vacancy({"name":"IT","description":"Super",
        ...                         "key_skills":"GOD]CoolBoy",
        ...                          "experience_id":"None",
        ...                          "premium":"Нет",
        ...                          "employer_name":"Газпром",
        ...                          "salary_gross":"Нет",
        ...                          "salary_from":"10",
        ...                          "salary_to":"20",
        ...                          "salary_currency":"RUR",
        ...                          "area_name":"Питер",
        ...                          "published_at":"2022-07-17T18:23:06+0300"}))
        Введите название файла: Введите название профессии: datetime.datetime(2022, 7, 17, 18, 23, 6)

        """
        city = x.area_name
        # 1 способ
        #time = x.published_at.split('-')
        #mili_time = time[2].split('T')[1].split('+')[0].split(':')
        # 2 способ - не эффективен
        #f = datetime.datetime.strptime(x.published_at,"%Y-%m-%dT%H:%M:%S+%f")
        # 3 способ
        year = x.published_at[0:4]

        try:
            f = currency_to_rub[x.salary.salary_currency]
        except:
            currency_to_rub[x.salary.salary_currency] = 0

        if not self.dict_inYear_WithName_and_city.__contains__(int(year)) and (x.name.__contains__(dataSet.job_name)
                                                                            or x.name.__contains__(
                    dataSet.job_name.lower()) and x.area_name == dataSet.city_name):
            self.dict_inYear_WithName_and_city[int(year)] = 1
            self.dict_inYear_WithName_salary_and_city[int(year)] = (float(x.salary.salary_from)
                                                                 * currency_to_rub[x.salary.salary_currency]
                                                                 + float(x.salary.salary_to)
                                                                 * currency_to_rub[x.salary.salary_currency]) \
                                                                / 2
        elif x.name.__contains__(dataSet.job_name) or x.name.__contains__(dataSet.job_name.lower()) and x.area_name == dataSet.city_name:
            self.dict_inYear_WithName_and_city[int(year)] += 1
            self.dict_inYear_WithName_salary_and_city[int(year)] += (float(x.salary.salary_from)
                                                                  * currency_to_rub[x.salary.salary_currency]
                                                                  + float(x.salary.salary_to)
                                                                  * currency_to_rub[x.salary.salary_currency]) \
                                                                 / 2
        if not self.dict_inYear_noName.__contains__(int(year)):
            self.dict_inYear_noName[int(year)] = 1
            self.dict_inYear_noName_salary[int(year)] = (float(x.salary.salary_from)
                                                               * currency_to_rub[x.salary.salary_currency]
                                                               + float(x.salary.salary_to)
                                                               * currency_to_rub[x.salary.salary_currency]) \
                                                              / 2
        elif self.dict_inYear_noName.__contains__(int(year)):
            self.dict_inYear_noName[int(year)] += 1
            self.dict_inYear_noName_salary[int(year)] += (float(x.salary.salary_from)
                                                                * currency_to_rub[x.salary.salary_currency]
                                                                + float(x.salary.salary_to)
                                                                * currency_to_rub[x.salary.salary_currency]) \
                                                               / 2
        if not self.temp_dict.__contains__(city):
            self.temp_dict[city] = 1
            self.temp_salary_dict[city] = (float(x.salary.salary_from)
                                              * currency_to_rub[x.salary.salary_currency]
                                              + float(x.salary.salary_to)
                                              * currency_to_rub[x.salary.salary_currency]) \
                                             / 2
        elif self.temp_dict.__contains__(city):
            self.temp_dict[city] += 1
            self.temp_salary_dict[city] += (float(x.salary.salary_from)
                                               * currency_to_rub[x.salary.salary_currency]
                                               + float(x.salary.salary_to)
                                               * currency_to_rub[x.salary.salary_currency]) \
                                              / 2
        return datetime.datetime(day=int(x.published_at[8:10]),
                                 month=int(x.published_at[5:7]),
                                 year=int(x.published_at[0:4]),
                                 hour=int(x.published_at[11:13]),
                                 minute=int(x.published_at[14:16]),
                                 second=int(x.published_at[17:19]))

    def get_year_sort(self, x):
        """
        Правило сортировки по опыту работу

        :param x (string) - строка с опытом работы в вакансии
        :return: string содержащий наибольший опыт работы для вакансии

        #>>> InputConect(DataSet()).get_year_sort("От 1 до 3 лет")
        Введите название файла: Введите название профессии: 3
        #>>> InputConect(DataSet()).get_year_sort("Нет")
        Введите название файла: Введите название профессии: 0
        #>>> InputConect(DataSet()).get_year_sort("От года до двух")
        Введите название файла: Введите название профессии: 0
        """
        split_list = x.split(" ")
        max_year = 0
        count_ind = 0
        for word in split_list:
            if word.isdigit():
                count_ind += 1
                max_year = max(max_year, int(word))
        if count_ind == 1:
            max_year += 1
        return max_year

    def get_sort_table(self, dataSet, table, full_skills):
        """
        Сортировка таблицы по любому из столбцов

        :param dataSet (DataSet) - Содержит в себе экземпляр обекта со всеми входными данными
        :param table (prettytable) - таблица для заполнения
        :param full_skills (list) - список с полными навыками
        :return: list отсортированных строк таблицы
        """
        column_for_sort = dataSet.sort_parameter
        index_column = list(table.field_names).index(column_for_sort)
        reversesort = dataSet.IsReverseSort
        if column_for_sort == "Навыки":
            sort_key = lambda x: len(self.find_full_skills(x, full_skills))
            return sorted(table.rows, key=sort_key, reverse=reversesort)
        elif column_for_sort == "Дата публикации вакансии":
            sort_key = lambda x: self.get_date_sort(x)
            return sorted(table.rows, key=sort_key, reverse=reversesort)
        elif column_for_sort == "Оклад":
            salar = Salary()
            sort_key = lambda x: salar.salary_sorter(x[index_column])
            return sorted(table.rows, key=sort_key, reverse=reversesort)
        elif column_for_sort == "Опыт работы":
            sort_key = lambda x: self.get_year_sort(x[index_column])
            return sorted(table.rows, key=sort_key, reverse=reversesort)
        elif column_for_sort != "":
            sort_key = lambda x: x[index_column]
            return sorted(table.rows, key=sort_key, reverse=reversesort)

    def get_sort_dataSet(self, dataSet):
        """
        Отсортировать список вакансий в экземпляре DataSet

        :param dataSet (DataSet) - Содержит в себе экземпляр обекта со всеми входными данными
        :return: DataSet с отсортированным списком вакансий
        """
        reversesort = dataSet.IsReverseSort
        sort_key = lambda x: self.get_date_sort(x)
        return sorted(dataSet.vacancies_objects, key=sort_key, reverse=reversesort)

    def sorted_for_graf(self):
        """
        Сортировка вакансий для составления графика. Производится сортировка по зарплате, вакансиям, их долям в городах и определенной профессии

        :return: InputConect с заполненными словарями вакансий, зарплат и их распределения по городам и определенной профессии
        """
        dataSet.sort_parameter = "Дата публикации вакансии"
        self.dataSet = self.get_sort_dataSet(self.dataSet)
        for key in self.dict_inYear_noName.keys():
            self.dict_inYear_noName_salary[key] = math.floor(
                int(self.dict_inYear_noName_salary[key]) / self.dict_inYear_noName[key])
            if len(self.dict_inYear_WithName_and_city) > 0:
                self.dict_inYear_WithName_salary_and_city[key] = math.floor(
                    int(self.dict_inYear_WithName_salary_and_city[key]) / self.dict_inYear_WithName_and_city[key])
            else:
                self.dict_inYear_WithName_and_city[key] = 0
                self.dict_inYear_WithName_salary_and_city[key] = 0
        bad_city_vac_count = 0
        for city in self.temp_dict.keys():
            if self.temp_dict[city] >= math.floor(len(dataSet.vacancies_objects) / 100):
                try:
                    self.dict_inYear_City[city] = round(int(self.temp_dict[city]) / len(dataSet.vacancies_objects), 4)
                    self.dict_inYear_City_salary[city] = math.floor(
                        int(self.temp_salary_dict[city]) / self.temp_dict[city])
                except:
                    f = 6
            else:
                bad_city_vac_count += int(self.temp_dict[city]) / len(dataSet.vacancies_objects)
        self.dict_inYear_City["Другие"] = bad_city_vac_count


class DataSet():
    """
    Класс содержащий данные вакансий и параметры работы с ними. Осуществляет чтение данных с файла и проверку их корректности

    Атрибуты:
        file_name (string) - содержит название файла csv с данными
        job_name (string) - содержит название профессии, по которой будет осуществлен поиск и отбор

        vacancies_objects (list) - содержит все поля вакансии
        title_piece (list) - список заголовков таблицы prettytable
        sort_parameter (string) - параметр сортировки таблицы prettytable
        IsReverseSort (bool) - определяет порядок сортировки (возрастание/убывание)
        filter_for_table (list) - список, содержащий элементы фильтра для таблицы prettytable
        vacant_piece (list) - список, содержащий вырезку вакансий из таблицы prettytable

        message_error (string) - текст сообщения об ошибке, возможной при проверке на корректность
        filter_atr (string) - полный запрос о фильтрации, требующий разбиения на составляющие и внесения в filter_for_table
        sort_atr (string) - полный запрос о сортировке, требующий проверки и внесения в sort_parameter
        revers_atr (string) - запрос о порядке сортировки, требующий проверки и внесения в IsReverseSort
    """
    file_name = ""
    job_name = ""
    city_name = ""

    vacancies_objects = []
    title_piece = ["№"] + list(vacant_dic.values())
    sort_parameter = ""
    IsReverseSort = False
    filter_for_table = []
    vacant_piece = []

    message_error = ""
    filter_atr = ""
    sort_atr = ""
    revers_atr = ""

    def __init__(self):
        """
        Инициализирует экземпляр объекта DataSet

        Args:
            file_name (string) - содержит название файла csv с данными
            job_name (string) - содержит название профессии, по которой будет осуществлен поиск и отбор
            vacancies_objects (list) - содержит все поля вакансии
        """
        self.file_name = input("Введите название файла: ")
        self.job_name = input("Введите название профессии: ")
        self.city_name = input("Введите название региона: ")

        # self.file_name = "vacancies_dif_currencies.csv"
        # self.job_name = "Аналитик"
        self.check_atr()
        self.vacancies_objects = self.csv_filter_pandas(self.file_name)
        if len(self.vacancies_objects) == 0:
            print("Нет данных")
            exit()


    def csv_filter_pandas(self,file):
        df = pd.read_csv(file, encoding='utf_8_sig')
        vacant_list = []
        for row in df.iterrows():
            vacant_list.append(Vacancy(row[1]))
        return vacant_list
    def csv_filter(self, file, data):
        """
        Фильтрация содержимых файла

        :param file (string) -  название файла csv с данными
        :param data (DataSet) -  DataSet в который будут занесены данные из файла
        :return: list полей вакансий
        """
        list_naming = []
        temp_read = self.csv_reader(file, data)
        for data in temp_read[1]:
            temp_dict = {}
            for i, row in enumerate(data):
                data[i] = data[i].replace('\n', ']')
                temp = re.split(r'<.*?>', data[i])
                if temp[0] == "True" or temp[0] == "TRUE":
                    temp[0] = "Да"
                elif temp[0] == "False" or temp[0] == "FALSE":
                    temp[0] = "Нет"
                data[i] = re.sub(r'\s+', ' ', ''.join(temp)).strip()
                descript = data[i].split("]")
                if len(descript) == 1:
                    descript = "".join(descript)
                else:
                    descript = "]".join(descript)
                temp_dict[temp_read[0][i]] = descript
            list_naming.append(Vacancy(temp_dict))
        return list_naming

    def csv_reader(self, file_name, data_self):
        """
        Чтение файла и проверка его на наличие данных

        :param file_name (string) - название файла csv с данными
        :param data_self (DataSet) -  DataSet в который будут занесены данные из файла
        :return: Touple содержащий строку с названием колонок и данными
        """
        title_row = []
        data_row = []
        if os.stat(file_name).st_size == 0:
            data_self.message_error = "Пустой файл"
            print(data_self.message_error)
            exit()
        with open(file_name, 'r', encoding='utf_8_sig') as f:
            reader = csv.reader(f)
            for index, row in enumerate(reader):
                if index == 0:
                    title_row = row
                    normal_length = len(row)
                else:
                    if '' in row and ((row[1] != '' or row[2] != '') and row[3] != ''):
                        temp = []
                        for item in row:
                            if item == '':
                                temp.append(item.replace('', '0'))
                            else:
                                temp.append(item)
                        row = temp
                    if normal_length == len(row) and '' not in row:
                        data_row.append(row)
        return title_row, data_row

    def check_atr(self):
        """
        Проверка атрибутов на корректность

        :return: Void
        """
        global vacant_dic
        if len(self.filter_atr) != 0:
            if not self.filter_atr.__contains__(':'):
                self.message_error = "Формат ввода некорректен"
                print(self.message_error)
                exit(self.message_error)
            parameter = self.filter_atr.split(":")[0]
            parameter_list = self.filter_atr.split(":")[1].split(", ")
            parameter_list[0] = parameter_list[0].strip()
            if not (list(vacant_dic.values()).__contains__(parameter) or parameter == "Идентификатор валюты оклада"):
                self.message_error = "Параметр поиска некорректен"
                print(self.message_error)
                exit(self.message_error)
            self.filter_for_table = self.filter_atr.split(":")
        if not list(vacant_dic.values()).__contains__(self.sort_atr) and not self.sort_atr == "":
            self.message_error = "Параметр сортировки некорректен"
            print(self.message_error)
            exit(self.message_error)
        else:
            self.sort_parameter = (lambda arg: arg if list(vacant_dic.values()).__contains__(arg) else "")(
                self.sort_atr)
        if not ["Да", "Нет", ""].__contains__(self.revers_atr):
            self.message_error = "Порядок сортировки задан некорректно"
            print(self.message_error)
            exit(self.message_error)
        else:
            self.IsReverseSort = (lambda arg: arg == "Да")(self.revers_atr)


class Vacancy():
    """
    Класс, содержащий поля вакансии для внесения их в список vacancies_objects в экземпляре DataSet

    Атрибуты:
        name (string) - название профессии
        description (string) - описание профессии
        key_skills (list) - навыки заявителя
        experience_id (string) - опыт работы заявителя
        premium (string) - показатель, является ли заявитель обладателем премиума
        employer_name (string) - название работодателя
        salary (Salary) - информация о зарплате заявителя
        area_name (string) - название города заявителя (проживания или работы)
        published_at (string) - дата публикации вакансии
    """
    name = ""
    description = ""
    key_skills = []
    experience_id = ""
    premium = ""
    employer_name = ""
    salary = ""
    area_name = ""
    published_at = ""

    def __init__(self, vacant):
        """
        Инициализирует экземпляр объекта Vacancy

        :param vacant (dict) - временный словарь для обработки считываемых значений вакансий
        Args:
            name (string) - название профессии
            description (string) - описание профессии
            key_skills (list) - навыки заявителя
            experience_id (string) - опыт работы заявителя
            premium (string) - показатель, является ли заявитель обладателем премиума
            employer_name (string) - название работодателя
            salary (Salary) - информация о зарплате заявителя
            area_name (string) - название города заявителя (проживания или работы)
            published_at (string) - дата публикации вакансии

        #>>> Vacancy({"name":"IT","description":"Super","key_skills":"GOD]CoolBoy","experience_id":"None","premium":"Нет","employer_name":"Газпром","salary_gross":"Нет","salary_from":"10","salary_to":"20","salary_currency":"RUR","area_name":"Питер","published_at":"30.12.2020"}).name
        'IT'
        #>>> Vacancy({"name":"IT","description":"Super","key_skills":"GOD]CoolBoy","experience_id":"None","premium":"Нет","employer_name":"Газпром","salary_gross":"Нет","salary_from":"10","salary_to":"20","salary_currency":"RUR","area_name":"Питер","published_at":"30.12.2020"}).description
        'Super'
        #>>> Vacancy({"name":"IT","description":"Super","key_skills":"GOD]CoolBoy","experience_id":"None","premium":"Нет","employer_name":"Газпром","salary_gross":"Нет","salary_from":"10","salary_to":"20","salary_currency":"RUR","area_name":"Питер","published_at":"30.12.2020"}).key_skills
        ['GOD', 'CoolBoy']
        #>>> Vacancy({"name":"IT","description":"Super","key_skills":"GOD]CoolBoy","experience_id":"None","premium":"Нет","employer_name":"Газпром","salary_gross":"Нет","salary_from":"10","salary_to":"20","salary_currency":"RUR","area_name":"Питер","published_at":"30.12.2020"}).experience_id
        'None'
        #>>> Vacancy({"name":"IT","description":"Super","key_skills":"GOD]CoolBoy","experience_id":"None","premium":"Нет","employer_name":"Газпром","salary_gross":"Нет","salary_from":"10","salary_to":"20","salary_currency":"RUR","area_name":"Питер","published_at":"30.12.2020"}).premium
        'Нет'
        #>>> Vacancy({"name":"IT","description":"Super","key_skills":"GOD]CoolBoy","experience_id":"None","premium":"Нет","employer_name":"Газпром","salary_gross":"Нет","salary_from":"10","salary_to":"20","salary_currency":"RUR","area_name":"Питер","published_at":"30.12.2020"}).employer_name
        'Газпром'
        #>>> Vacancy({"name":"IT","description":"Super","key_skills":"GOD]CoolBoy","experience_id":"None","premium":"Нет","employer_name":"Газпром","salary_gross":"Нет","salary_from":"10","salary_to":"20","salary_currency":"RUR","area_name":"Питер","published_at":"30.12.2020"}).area_name
        'Питер'
        #>>> Vacancy({"name":"IT","description":"Super","key_skills":"GOD]CoolBoy","experience_id":"None","premium":"Нет","employer_name":"Газпром","salary_gross":"Нет","salary_from":"10","salary_to":"20","salary_currency":"RUR","area_name":"Питер","published_at":"30.12.2020"}).published_at
        '30.12.2020'
        """
        self.name = vacant["name"]
        gross = "None"
        try:

            self.description = vacant["description"]
            self.key_skills = vacant["key_skills"].split(']')
            self.experience_id = vacant["experience_id"]
            self.premium = vacant["premium"]
            self.employer_name = vacant["employer_name"]
            gross = vacant["salary_gross"]

        except:
            f = 5
        self.salary = Salary(vacant["salary_from"], vacant["salary_to"], gross,
                             vacant["salary_currency"])
        self.area_name = vacant["area_name"]
        self.published_at = vacant["published_at"]


class Salary():
    """
    Класс для представления данных о зарплате заявителя

    Атрибуты:
        salary_from (string) - минимальная зарплата
        salary_to (string) - максимальная зарплата
        salary_gross (string) - налоговый вычет заложен в зарплату
        salary_currency (string) - валюта зарплата

    """
    salary_from = ""
    salary_to = ""
    salary_gross = ""
    salary_currency = ""

    def __init__(self, *args):
        """
        Инициализация экземпляра Salary с проверкой salary_gross на наличие значения

        :param args (list) - со значениями полей зарплаты из вакансии

        Args:
            salary_from (string) - минимальная зарплата
            salary_to (string) - максимальная зарплата
            salary_gross (string) - налоговый вычет заложен в зарплату
            salary_currency (string) - валюта зарплата

        #>>> Salary("300","500","Нет","USD").salary_from
        '300'
        #>>> Salary("300","500","Нет","USD").salary_to
        '500'
        #>>> Salary("300","500","Нет","USD").salary_currency
        'USD'
        """
        global curr_dict
        if len(args) > 0:
            self.salary_from = args[0]
            self.salary_to = args[1]
            if args[2] != "None":
                self.salary_gross = args[2]
            self.salary_currency = args[3]
            if not curr_dict.__contains__(args[3]):
                curr_dict[args[3]] = 1
            else:
                curr_dict[args[3]] += 1

    def prepare_salary(self, string_salary):
        """
        Перевод зарплаты в нужный вид для форматирования

        :param string_salary (string) - строка со значением всех полей Salary
        :return: string с нужным значением зарплаты (минимальным или максимальным)
        """
        list_numb = []
        count = 0
        for char in reversed(string_salary):
            if count < 3:
                list_numb.append(char)
                count += 1
            else:
                list_numb.append(" ")
                list_numb.append(char)
                count = 0
        return "".join(list_numb.__reversed__())

    def salary_sorter(self, x):
        """
        Сортировка по окладу (среднему)

        :param x (Salary) - экземпляр объекта Salary, который требует сортировки
        :return: float среднего значения зарплаты для определенной вакансии

        #>>> Salary().salary_sorter(Salary(30,50,"None","RUR"))
        40.0
        #>>> Salary().salary_sorter(Salary(30,50,"None","UZS"))
        0.21999999999999997
        #>>> Salary().salary_sorter(Salary(0,0,"Yes","USD"))
        0.0
        """
        currency = x.salary_currency
        salar_min = float(x.salary_from) * currency_to_rub[currency]
        salar_max = float(x.salary_to) * currency_to_rub[currency]
        return (salar_min + salar_max) / 2


class report():
    """
    Класс для составления отчета по вакансиям

    Атрибуты:
        # Характеристики таблицы .xlsx
        border (Border) - значения границ таблицы .xlsx
        font (Font) - значения шрифта таблицы .xlsx

        # Наполнение
        total_year (list) - лист с годами подачи вакансий
        mean_salary (dict) - словарь со средними значениями зарплат
        mean_salary_job (dict) - словарь со средними значениями зарплат для определенной профессии
        count_vac (dict) - словарь всех вакансий
        count_vac_job (dict) - словарь всех вакансий для определенной профессии

        mean_salary_city (dict) - словарь со средними значениями зарплат для определенного города
        count_vac_city - словарь всех вакансий для определенного города

        book (Workbook) - экземпляр Workbook для создания таблиц в .xlsx файле

    """
    border = 0
    font = 0

    total_year = []
    mean_salary = {}
    mean_salary_job = {}
    count_vac = {}
    count_vac_job = {}

    mean_salary_city = {}
    count_vac_city = {}

    book = 0

    def __init__(self, font, border):
        """
        Инициализация экземпляра объекта report

        :param font (Font) - значения шрифта таблицы .xlsx
        :param border (Border) - значения границ таблицы .xlsx
        """
        self.font = font
        self.border = border

    def generate_excel(self, years, data, book, name_job):
        """
        Создание excel таблицы формата .xlsx

        :param years (list) - лист с годами подачи вакансий
        :param data (DataSet) - Содержит в себе экземпляр обекта со всеми входными данными
        :param book (Workbook) - экземпляр Workbook для создания таблиц в .xlsx файле
        :param name_job (string) - название профессии
        :return: .xlsx файл "report.xlsx"
        """

        self.total_year = years
        self.mean_salary = data[0]
        self.mean_salary_job = data[1]
        self.count_vac = data[2]
        self.count_vac_job = data[3]
        self.mean_salary_city = data[4]
        self.count_vac_city = data[5]

        # 1
        ws = book.active
        ws.title = "Статистика по годам"
        book.create_sheet("Статистика по городам", 1)
        ws['A1'] = "Год"
        ws['A1'].font = self.font

        ws['B1'] = "Средняя зарплата"
        ws['B1'].font = self.font

        ws['C1'] = "Средняя зарплата - {0}".format(name_job)
        ws['C1'].font = self.font

        ws['D1'] = "Количество вакансий"
        ws['D1'].font = self.font

        ws['E1'] = "Количество вакансий - {0}".format(name_job)
        ws['E1'].font = self.font

        for i in range(len(years)):
            ws['A{0}'.format(i + 2)] = years[i]
            ws['B{0}'.format(i + 2)] = self.mean_salary[years[i]]
            ws['C{0}'.format(i + 2)] = self.mean_salary_job[years[i]]
            ws['D{0}'.format(i + 2)] = self.count_vac[years[i]]
            ws['E{0}'.format(i + 2)] = self.count_vac_job[years[i]]

        column_widths = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        for row in ws.rows:
            for i, cell in enumerate(row):
                column_widths[i + 1] = max(len((str)(cell.value)) + 1, column_widths[i + 1])
                cell.border = self.border
        for i in range(len(column_widths)):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_widths[i + 1]

        # 2
        ws = book["Статистика по городам"]
        ws['A1'] = "Город"
        ws['A1'].font = self.font

        ws['B1'] = "Уровень зарплат"
        ws['B1'].font = self.font

        ws['D1'] = "Город"
        ws['D1'].font = self.font

        ws['E1'] = "Доля вакансий".format(name_job)
        ws['E1'].font = self.font

        cityes_salar = list(self.mean_salary_city.keys())
        cityes_vac = list(self.count_vac_city.keys())
        for i in range(len(cityes_salar)):
            ws['A{0}'.format(i + 2)] = cityes_salar[i]
            ws['B{0}'.format(i + 2)] = self.mean_salary_city[cityes_salar[i]]
            ws['D{0}'.format(i + 2)] = cityes_vac[i]
            ws['E{0}'.format(i + 2)] = self.count_vac_city[cityes_vac[i]]
            ws['E{0}'.format(i + 2)].number_format = "0%"

        column_widths = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        for row in ws.rows:
            for i, cell in enumerate(row):
                column_widths[i + 1] = max(len((str)(cell.value)) + 1, column_widths[i + 1])
                cell.border = self.border
        for i in range(len(column_widths)):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_widths[i + 1]

    def generate_excel_async(self, years, data, name_job, book):

        self.total_year = years
        self.mean_salary = data[0]
        self.mean_salary_job = data[1]
        self.count_vac = data[2]
        self.count_vac_job = data[3]
        self.mean_salary_city = data[4]
        self.count_vac_city = data[5]

        for year in years:
            index = years.index(year)
            header = ["Год", "Средняя зарплата", "Средняя зарплата - {0}".format(name_job), "Количество вакансий",
                      "Количество вакансий - {0}".format(name_job)]
            data = {"Год": year, "Средняя зарплата": self.mean_salary[years[index]],
                    "Средняя зарплата - {0}".format(name_job): self.mean_salary_job[years[index]],
                    "Количество вакансий": self.count_vac[years[index]],
                    "Количество вакансий - {0}".format(name_job): self.count_vac_job[years[index]]}

            file = pd.DataFrame(data, columns=header, index=[0])
            file.to_csv("{0}/{1}.csv".format("years", year))
    def generate_report(self,data):
        """
        Создание файла .pdf с графиками и таблицами полученным в результате формирования экземпляра объекта report

        :param data (DataSet) - Содержит в себе экземпляр обекта со всеми входными данными
        :return: .pdf файл "report.pdf"
        """
        # В PDF
        job_name = data.job_name
        city_name = data.city_name

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("Shablon.html")
        options = {
            'encoding': "UTF-8",
            'enable-local-file-access': None,
            'no-outline': None
        }

        pdf_template = template.render(
            {'job_name': job_name, 'city_name' : city_name,'excel': wb["Статистика по годам"], 'excel2': wb['Статистика по городам']})
        config = pdfkit.configuration(wkhtmltopdf=r'E:\Program Files (x86)\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)


class Salary_sorter_tests(TestCase):
    def test_RU_salary(self):
        self.assertEqual(Salary().salary_sorter(Salary(30,50,"None","RUR")),40.0)
    def test_Not_RU_salary(self):
        self.assertEqual(Salary().salary_sorter(Salary(30,50,"None","UZS")),0.21999999999999997)
    def test_Bad_salary(self):
        self.assertEqual(Salary().salary_sorter(Salary(0,0,"Yes","USD")),0.0)

class Salary_init_tests(TestCase):
    def test_salary_from(self):
        self.assertEqual(Salary("300","500","Нет","USD").salary_from,"300")
    def test_salary_to(self):
        self.assertEqual(Salary("300","500","Нет","USD").salary_to,"500")
    def test_salary_currency(self):
        self.assertEqual(Salary("300","500","Нет","USD").salary_currency,"USD")

class Vacancy_tests(TestCase):
    def test_name(self):
        self.assertEqual(Vacancy({"name":"IT","description":"Super",
                                  "key_skills":"GOD]CoolBoy",
                                  "experience_id":"None",
                                  "premium":"Нет",
                                  "employer_name":"Газпром",
                                  "salary_gross":"Нет",
                                  "salary_from":"10",
                                  "salary_to":"20",
                                  "salary_currency":"RUR",
                                  "area_name":"Питер",
                                  "published_at":"30.12.2020"}).name,
                         "IT")
    def test_description(self):
        self.assertEqual(Vacancy(
            {"name": "IT", "description": "Super", "key_skills": "GOD]CoolBoy", "experience_id": "None",
             "premium": "Нет", "employer_name": "Газпром", "salary_gross": "Нет", "salary_from": "10",
             "salary_to": "20", "salary_currency": "RUR", "area_name": "Питер", "published_at": "30.12.2020"}).description,
                         "Super")
    def test_key_skills(self):
        self.assertEqual(Vacancy(
            {"name": "IT", "description": "Super", "key_skills": "GOD]CoolBoy", "experience_id": "None",
             "premium": "Нет", "employer_name": "Газпром", "salary_gross": "Нет", "salary_from": "10",
             "salary_to": "20", "salary_currency": "RUR", "area_name": "Питер",
             "published_at": "30.12.2020"}).key_skills,
                         ['GOD','CoolBoy'])
    def test_expe(self):
        self.assertEqual(Vacancy(
            {"name": "IT", "description": "Super", "key_skills": "GOD]CoolBoy", "experience_id": "None",
             "premium": "Нет", "employer_name": "Газпром", "salary_gross": "Нет", "salary_from": "10",
             "salary_to": "20", "salary_currency": "RUR", "area_name": "Питер",
             "published_at": "30.12.2020"}).experience_id,
                         'None')
    def test_premium(self):
        self.assertEqual(Vacancy(
            {"name": "IT", "description": "Super", "key_skills": "GOD]CoolBoy", "experience_id": "None",
             "premium": "Нет", "employer_name": "Газпром", "salary_gross": "Нет", "salary_from": "10",
             "salary_to": "20", "salary_currency": "RUR", "area_name": "Питер",
             "published_at": "30.12.2020"}).premium,
                         'Нет')
    def test_eployer(self):
        self.assertEqual(Vacancy(
            {"name": "IT", "description": "Super", "key_skills": "GOD]CoolBoy", "experience_id": "None",
             "premium": "Нет", "employer_name": "Газпром", "salary_gross": "Нет", "salary_from": "10",
             "salary_to": "20", "salary_currency": "RUR", "area_name": "Питер",
             "published_at": "30.12.2020"}).employer_name,
                         'Газпром')
    def test_area_name(self):
        self.assertEqual(Vacancy(
            {"name": "IT", "description": "Super", "key_skills": "GOD]CoolBoy", "experience_id": "None",
             "premium": "Нет", "employer_name": "Газпром", "salary_gross": "Нет", "salary_from": "10",
             "salary_to": "20", "salary_currency": "RUR", "area_name": "Питер",
             "published_at": "30.12.2020"}).area_name,
                         'Питер')
    def test_date(self):
        self.assertEqual(Vacancy(
            {"name": "IT", "description": "Super", "key_skills": "GOD]CoolBoy", "experience_id": "None",
             "premium": "Нет", "employer_name": "Газпром", "salary_gross": "Нет", "salary_from": "10",
             "salary_to": "20", "salary_currency": "RUR", "area_name": "Питер",
             "published_at": "30.12.2020"}).published_at,
                         '30.12.2020')

class sort_date_test(TestCase):
    def test_begin(self):
        self.assertEqual(InputConect(DataSet()).get_date_sort(Vacancy({"name":"IT","description":"Super",
                       "key_skills":"GOD]CoolBoy",
                        "experience_id":"None",
                        "premium":"Нет",
                      "employer_name":"Газпром",
                   "salary_gross":"Нет",
                      "salary_from":"10",
                        "salary_to":"20",
                       "salary_currency":"RUR",
                       "area_name":"Питер",
                       "published_at":'2022-07-17T18:23:06+0300'})),datetime.datetime(2022, 7, 17, 18, 23, 6))

class sort_year_tests(TestCase):
    def test_normal(self):
        self.assertEqual(InputConect(DataSet()).get_year_sort("От 1 до 3 лет"),3)
    def test_zero(self):
        self.assertEqual(InputConect(DataSet()).get_year_sort("Нет"),
                         0)
    def test_bad(self):
        self.assertEqual(InputConect(DataSet()).get_year_sort("От года до двух"),
                         0)

dataSet = DataSet()

# table = prettytable.PrettyTable()
# table.hrules = prettytable.ALL
# table.align = 'l'
full_table_skills = {}
full_table_date = {}


# Динамика зарплат по годам
sorter_master = InputConect(dataSet)
# Сложная функция
#sorter_master.sorted_for_graf()
# temp_dict = []
# temp_vacant_obj = []
# for curr in curr_dict.items():
#     curr_dict[curr[0]] = curr[1] / len(dataSet.vacancies_objects)
#     if curr[1] > 5000:
#         temp_dict.append(curr[0])
#
#
# for vacant in dataSet.vacancies_objects:
#     if temp_dict.__contains__(vacant.salary.salary_currency):
#         temp_vacant_obj.append(vacant)
#
# dataSet.vacancies_objects = temp_vacant_obj
# dataSet.vacancies_objects.sort(key= lambda x: x.published_at)
# min = dataSet.vacancies_objects[0].published_at.split('T')[0]
# max = dataSet.vacancies_objects[len(dataSet.vacancies_objects) - 1].published_at.split('T')[0]
# minTime = datetime.datetime(year=int(min.split('-')[0]),
#                             month=int(min.split('-')[1]),
#                             day=1)
# maxTime = datetime.datetime(year=int(max.split('-')[0]),
#                             month=int(max.split('-')[1]),
#                             day=1)
#
# years_dif = maxTime.year - minTime.year
# month_dif = maxTime.month - minTime.month + (12 * years_dif)

# Создание csv валют
# header_csv = ["date"] + list(curr_dict.keys())
# file = open('current.csv', 'w', newline='')
# writer = csv.DictWriter(file,fieldnames=header_csv)
# writer.writeheader()
#
# row_list = []
# df = pd.DataFrame(columns=header_csv)
# month = minTime.month
# year = minTime.year
# for date in range(month_dif):
#     answ_dict = {"date": 0}
#     dat = ""
#     for curr in curr_dict.keys():
#         answ_dict["date"] = "{0}-{1}".format(year, month)
#         if curr != "RUR":
#             m_month = month
#             if month < 10:
#                 m_month = "0{0}".format(month)
#             url_Bank = "http://www.cbr.ru/scripts/XML_daily.asp?date_req=01/{0}/{1}".format(m_month, year)
#             try:
#                 res = ET.fromstring(requests.get(url_Bank).text) \
#                     .find('./Valute[CharCode="{0}"]/Value'.format(curr)) \
#                     .text.replace(',', '.')
#                 answ_dict[curr] = res
#             except:
#                 answ_dict[curr] = ' '
#         else:
#             answ_dict[curr] = ' '
#     month += 1
#     if month > 12:
#         month = 1
#         year += 1
#     writer.writerow(answ_dict)
#     answ_dict.clear()
#
# file.close()
#print("Частота валют за промежуток 2003-2022 : {0}".format(curr_dict))

# Попытка мультипроцессинка №64-2
# if __name__ == "__main__":
#     p1 = multiprocessing.Process(target=pool_work, args=(x,))
#     p1.start()
#
#     p2 = multiprocessing.Process(target=pool_work2, args=(x,))
#     p2.start()
#
#     p3 = multiprocessing.Process(target=pool_work3, args=(x,))
#     p3.start()
#
#     p1.join()
#     p2.join()
#     p3.join()

sorter_master.dict_inYear_City_salary = dict(
    sorted(sorter_master.dict_inYear_City_salary.items(), key=lambda item: item[1], reverse=True))
sorter_master.dict_inYear_City = dict(
    sorted(sorter_master.dict_inYear_City.items(), key=lambda item: item[1], reverse=True))

sumInList = sorter_master.dict_inYear_City["Другие"] + sum(list(dict(list(sorter_master.dict_inYear_City.items())[10:]).values()))
sorter_master.dict_inYear_City["Другие"] = 0
sorter_master.dict_inYear_City = dict(
    sorted(sorter_master.dict_inYear_City.items(), key=lambda item: item[1], reverse=True))
sorter_master.dict_inYear_City = dict(list(sorter_master.dict_inYear_City.items())[:10])
sorter_master.dict_inYear_City_salary = dict(list(sorter_master.dict_inYear_City_salary.items())[:10])

print("Динамика уровня зарплат по годам: {0}".format(sorter_master.dict_inYear_noName_salary))
print("Динамика количества вакансий по годам: {0}".format(sorter_master.dict_inYear_noName))
print("Динамика уровня зарплат по годам для выбранной профессии и региона: {0}".format(sorter_master.dict_inYear_WithName_salary_and_city))
print("Динамика количества вакансий по годам для выбранной профессии и региона: {0}".format(sorter_master.dict_inYear_WithName_and_city))
# print("Уровень зарплат по городам (в порядке убывания): {0}".format(sorter_master.dict_inYear_City_salary))
# print("Доля вакансий по городам (в порядке убывания): {0}".format(sorter_master.dict_inYear_City))
font_title = Font(name='Calibri',
                  size=11,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

wb = op.Workbook()

#Таблица
rep = report(font_title,border)
data_for_excel = [sorter_master.dict_inYear_noName_salary,
           sorter_master.dict_inYear_WithName_salary_and_city,
            sorter_master.dict_inYear_noName,
            sorter_master.dict_inYear_WithName_and_city,
            sorter_master.dict_inYear_City_salary,
            sorter_master.dict_inYear_City]
rep.generate_excel_async(list(sorter_master.dict_inYear_noName_salary.keys()),data_for_excel,dataSet.job_name,wb)
rep.generate_excel(list(sorter_master.dict_inYear_noName_salary.keys()),data_for_excel,wb,dataSet.job_name)
wb.save("report.xlsx")
#Графики
labels_years = list(sorter_master.dict_inYear_noName_salary.keys())
salary_noName = list(sorter_master.dict_inYear_noName_salary.values())
salart_Name = list(sorter_master.dict_inYear_WithName_salary_and_city.values())

vac_noName = list(sorter_master.dict_inYear_noName.values())
vac_Name = list(sorter_master.dict_inYear_WithName_and_city.values())

cityes_salary = list(sorter_master.dict_inYear_City_salary.values())
labels_cityes = list(sorter_master.dict_inYear_City.keys())

sorter_master.dict_inYear_City["Другие"] = sumInList
sorter_master.dict_inYear_City = dict(
    sorted(sorter_master.dict_inYear_City.items(), key=lambda item: item[1], reverse=True))
circle_labels = list(sorter_master.dict_inYear_City.keys())
cityes_perc = list(sorter_master.dict_inYear_City.values())

width = 0.4
x = np.arange(len(labels_years))
y = np.arange(len(labels_cityes))

matplotlib.rc('axes', titlesize=8)
matplotlib.rc('font', size=8)
matplotlib.rc('xtick', labelsize=8)
matplotlib.rc('ytick', labelsize=8)
matplotlib.rc('legend', fontsize=8)

fig, ax = plt.subplots(2, 2)

rects1 = ax[0, 0].bar(x - width / 2, salary_noName, width, label="Средняя з/п")
rects2 = ax[0, 0].bar(x + width / 2, salart_Name, width, label="з/п {0}({1})".format(dataSet.job_name, dataSet.city_name))
ax[0, 0].set_title('Уровень зарплат по годам')
ax[0, 0].set_xticks(x)
ax[0, 0].set_xticklabels(labels_years, rotation=90)
ax[0, 0].legend()

rects3 = ax[0, 1].bar(x - width / 2, vac_noName, width, label="Количество вакансий")
rects4 = ax[0, 1].bar(x + width / 2, vac_Name, width, label="Количество вакансий {0}({1})".format(dataSet.job_name, dataSet.city_name))
ax[0, 1].set_title('Количество вакансий по годам')
ax[0, 1].set_xticks(x)
ax[0, 1].set_xticklabels(labels_years, rotation=90)
ax[0, 1].legend()

rects5 = ax[1, 0].barh(y, cityes_salary, width * 2, align='center')
ax[1, 0].set_title('Уровень зарплат по городам')
ax[1, 0].set_yticks(y, labels=labels_cityes)
ax[1, 0].set_yticklabels(labels_cityes, fontsize=6,
                         fontdict={'horizontalalignment': 'right', 'verticalalignment': 'center'})
ax[1, 0].invert_yaxis()

circle = ax[1, 1].pie(cityes_perc, labels=circle_labels, textprops={'fontsize': 6})
ax[1, 1].set_title('Доля вакансий по городам', fontsize=6)
ax[1, 1].axis('equal')

plt.tight_layout()
fig.savefig("graph.png")

rep.generate_report(dataSet)
#
#rep.generate_report(dataSet)
#if __name__ == '__main__':
#    unittest.main()

# def get_curr(row,df_curr):
#     time = row['published_at'].split('T')[0]
#     year = time.split('-')[0]
#     month = time.split('-')[1]
#     if month[0] == '0':
#         month = month[1]
#     try:
#         curr = df_curr.loc[df_curr['date'] == "{0}-{1}".format(year,month)]
#         answer = curr[row['salary_currency']][0]
#     except:
#         answer = np.nan
#     return float(answer)
#
# df = pd.read_csv("vacancies_dif_currencies.csv", encoding='utf_8_sig')
# df_curr = pd.read_csv('current.csv')
# salary_column = []
# for row in df.iterrows():
#     if pd.isna(row[1]['salary_to']):
#         row[1]['salary_to'] = 0
#     elif pd.isna(row[1]['salary_from']):
#         row[1]['salary_from'] = 0
#     if pd.isna(row[1]['salary_currency']):
#         salary_column.append(row[1]['salary_currency'])
#     else:
#         salary_column.append((float(row[1]['salary_to']) + float(row[1]['salary_from'])) * get_curr(row[1],df_curr))
#
#
# df = df.replace({'salary_from':salary_column})
# df = df.drop(['salary_to','salary_currency'], axis=1)
# df = df.rename(columns={'salary_from':'salary'})
# df.head(100).to_csv('head100.csv')